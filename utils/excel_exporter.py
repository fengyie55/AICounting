import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime

class ExcelExporter:
    def __init__(self):
        pass
    
    def export_to_excel(self, records, file_path):
        """
        导出计数记录到Excel
        :param records: 计数记录列表
        :param file_path: 保存路径
        """
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "计数记录"
        
        # 设置表头
        headers = ["序号", "时间", "目标ID", "方向", "类别", "班次"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # 设置表头样式
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 填充数据
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=record['datetime'])
            ws.cell(row=row, column=3, value=record['track_id'])
            ws.cell(row=row, column=4, value="向上" if record['direction'] == 'up' else "向下")
            ws.cell(row=row, column=5, value=record['class'])
            ws.cell(row=row, column=6, value=self._shift_to_str(record['shift']))
        
        # 设置列宽
        column_widths = [8, 20, 10, 10, 10, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 添加统计工作表
        ws_stats = wb.create_sheet(title="统计分析")
        
        # 统计信息
        total_count = len(records)
        count_up = sum(1 for r in records if r['direction'] == 'up')
        count_down = sum(1 for r in records if r['direction'] == 'down')
        
        # 按班次统计
        shift_counts = {}
        for r in records:
            shift = r['shift']
            shift_counts[shift] = shift_counts.get(shift, 0) + 1
        
        # 填充统计数据
        stats_data = [
            ["统计项", "数值"],
            ["总计数", total_count],
            ["向上计数", count_up],
            ["向下计数", count_down],
            ["早班计数", shift_counts.get(0, 0)],
            ["中班计数", shift_counts.get(1, 0)],
            ["晚班计数", shift_counts.get(2, 0)],
            ["导出时间", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for row, data in enumerate(stats_data, 1):
            for col, value in enumerate(data, 1):
                cell = ws_stats.cell(row=row, column=col, value=value)
                if row == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 设置统计列宽
        ws_stats.column_dimensions['A'].width = 15
        ws_stats.column_dimensions['B'].width = 25
        
        # 保存文件
        wb.save(file_path)
    
    def _shift_to_str(self, shift):
        """
        班次转换为字符串
        """
        if shift == 0:
            return "早班"
        elif shift == 1:
            return "中班"
        elif shift == 2:
            return "晚班"
        else:
            return "未知"
    
    def export_daily_report(self, records, file_path):
        """
        导出日报表
        """
        # 按日期分组
        daily_data = {}
        for record in records:
            date = record['datetime'].split(' ')[0]
            if date not in daily_data:
                daily_data[date] = {
                    'total': 0,
                    'up': 0,
                    'down': 0,
                    'shifts': [0, 0, 0]
                }
            
            daily_data[date]['total'] += 1
            if record['direction'] == 'up':
                daily_data[date]['up'] += 1
            else:
                daily_data[date]['down'] += 1
            
            daily_data[date]['shifts'][record['shift']] += 1
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "日报表"
        
        # 表头
        headers = ["日期", "总计数", "向上", "向下", "早班", "中班", "晚班"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 填充数据
        row = 2
        for date in sorted(daily_data.keys()):
            data = daily_data[date]
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=data['total'])
            ws.cell(row=row, column=3, value=data['up'])
            ws.cell(row=row, column=4, value=data['down'])
            ws.cell(row=row, column=5, value=data['shifts'][0])
            ws.cell(row=row, column=6, value=data['shifts'][1])
            ws.cell(row=row, column=7, value=data['shifts'][2])
            row += 1
        
        # 设置列宽
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 保存
        wb.save(file_path)
